# -*- coding: utf-8 -*-
"""
01_univariate_cleaning.py
基于多源数据的大学生行为分析系统 - 单变量异常值清洗

数学原理与清洗策略：
1. 业务规则绝对边界截断 (Hard Boundary Clipping)
   - 成绩列：值 ∉ [0,100] 或 [0,150] → 置为 NaN
   - 单日时长列：值 ∉ [0,24] → 截断至边界

2. 盖帽法 (Winsorization)
   - 正态分布：Z = (x - μ) / σ，|Z| > 3 截断至 μ±3σ（约 99.7% 置信）
   - 偏态分布：IQR = Q3 - Q1，边界 Q1-1.5*IQR 与 Q3+1.5*IQR，超出截断

3. 缺失值填补：连续型用中位数，分类型用众数
"""

import shutil
import sys
import warnings
from pathlib import Path
from typing import Optional, Tuple

warnings.filterwarnings("ignore")
import numpy as np
import pandas as pd
from scipy import stats

# ==================== 配置 ====================
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "excel type"
FIELD_DESC_DIR = BASE_DIR / "doc type"
DATA_FALLBACK = BASE_DIR / "excel type"
OUTPUT_DIR = BASE_DIR / "cleaned_data_v1"

# 学号列（不参与盖帽法，保持原样）
ID_COL_NAMES = {"student_id", "XH", "XSBH", "LOGIN_NAME", "CREATER_LOGIN_NAME", "cardld", "IDSERTAL", "USERNUM", "KS_XH", "SID", "XHHGH"}
# 学号列映射（表名 -> 学号列名，供 load_field_descriptions 参考）
STUDENT_ID_COLS = {
    "上网统计": "XSBH", "体测数据": "XH", "体育课": "XH", "四六级成绩": "KS_XH",
    "图书馆打卡记录": "cardld", "奖学金获奖": "XSBH", "学生体能考核": "XH",
    "学生作业提交记录": "CREATER_LOGIN_NAME", "学生基本信息": "XH", "学生成绩": "XH",
    "学生签到记录": "LOGIN_NAME", "学生选课信息": "XH", "学科竞赛": "XHHGH",
    "学籍异动": "XH", "日常锻炼": "XH", "本科生综合测评": "XH", "毕业去向": "SID",
    "社团活动": "XSBH", "线上学习（综合表现）": "LOGIN_NAME", "考勤-汇总表": "XH",
    "考试提交记录": "CREATER_LOGIN_NAME", "讨论记录": "CREATER_LOGIN_NAME",
    "课堂任务参与": "LOGIN_NAME", "课程信息": None, "跑步打卡": "USERNUM",
    "门禁数据": "IDSERTAL", "上课信息统计表": None,
}

# 成绩/分数列关键词（用于硬边界 [0,100] 或 [0,150]）
SCORE_KEYWORDS = ["成绩", "分数", "score", "cj", "CJ", "绩点", "gpa", "JDCJ", "KCCJ", "ZF", "CFBFS", "SJCJ", "HSCJ", "BFB"]
# 单日时长列关键词（单位：小时，限制 [0, 24]）
DURATION_KEYWORDS = ["时长", "时间", "duration", "hours", "SWLJSC", "VIDEOJOB_TIME", "COURSE_LIVE_TIME", "SPECIAL_TIME"]


def load_field_descriptions() -> dict:
    """从 字段说明/ 解析学号列映射。"""
    mapping = {}
    try:
        from docx import Document
    except ImportError:
        return mapping
    if not FIELD_DESC_DIR.exists():
        return mapping
    keywords = ["学号", "学生编号", "账号", "登录名", "cardId", "cardld"]
    for doc_path in FIELD_DESC_DIR.glob("*.docx"):
        try:
            doc = Document(doc_path)
            tname = doc_path.stem
            for table in doc.tables:
                for row in table.rows:
                    cells = [c.text.strip() for c in row.cells]
                    if len(cells) >= 2:
                        pname, desc = cells[0], cells[1] if len(cells) > 1 else ""
                        if any(k in desc for k in keywords) or pname in ["XH", "XSBH", "LOGIN_NAME", "CREATER_LOGIN_NAME", "cardld", "IDSERTAL", "USERNUM", "KS_XH", "SID", "XHHGH"]:
                            mapping[tname] = pname
                            break
        except Exception:
            pass
    return mapping


def load_excel(p: Path) -> Optional[pd.DataFrame]:
    """
    多引擎尝试读取：openpyxl -> 默认 -> calamine。
    对损坏的 xlsx（BadZipFile），尝试 openpyxl read_only+data_only 模式。
    """
    # 常规读取
    for engine in ["openpyxl", None, "calamine"]:
        try:
            return pd.read_excel(p, engine=engine) if engine else pd.read_excel(p)
        except Exception:
            pass
    # 尝试 openpyxl 手动读取（read_only 模式有时能绕过部分损坏）
    try:
        from openpyxl import load_workbook
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if rows:
            df = pd.DataFrame(rows[1:], columns=rows[0])
            return df
    except Exception:
        pass
    return None


def ensure_student_id(df: pd.DataFrame, tname: str, field_mapping: dict) -> pd.DataFrame:
    """任务1：解析与对齐。字段保持原样，不重命名为 student_id。"""
    return df


def is_score_column(col: str) -> bool:
    return any(k in str(col) for k in SCORE_KEYWORDS)


def is_duration_column(col: str) -> bool:
    return any(k in str(col) for k in DURATION_KEYWORDS)


def hard_boundary_clipping(df: pd.DataFrame) -> pd.DataFrame:
    """
    任务2：业务规则绝对边界截断 (Hard Boundary Clipping)。
    数学原理：基于物理与教育常识的硬性约束，非统计推断。
    - 成绩列：值 ∉ [0,100]∪[0,150] → 置为 NaN（后续填补）
    - 单日时长列（小时）：值 ∉ [0,24] → clip(lower=0, upper=24)
    """
    df = df.copy()
    for col in df.select_dtypes(include=[np.number]).columns:
        s = df[col]
        if is_score_column(col):
            # 成绩：通常 [0,100]，四六级等可能 [0,150]
            invalid = (s < 0) | (s > 150)
            if invalid.any():
                df.loc[invalid, col] = np.nan
        elif is_duration_column(col):
            # 单日时长：物理上 [0, 24] 小时
            df[col] = s.clip(lower=0, upper=24)
    return df


def winsorize_column(s: pd.Series, method: str = "auto") -> Tuple[pd.Series, int]:
    """
    任务3：盖帽法 (Winsorization)。
    数学原理：
    - 正态分布：Z = (x - μ) / σ，|Z| > 3 时 P(|Z|>3) < 0.3%，截断至 μ±3σ
    - 偏态分布：IQR = Q3 - Q1，下界 L = Q1 - 1.5*IQR，上界 U = Q3 + 1.5*IQR，超出截断至 L/U
    返回 (处理后的 Series, 修正的异常值个数)
    """
    s = s.copy()
    valid = s.dropna()
    if len(valid) < 10 or valid.nunique() < 2:
        return s, 0
    if method == "auto":
        try:
            skew_val = float(stats.skew(valid))
            method = "zscore" if abs(skew_val) < 0.5 else "iqr"
        except Exception:
            method = "iqr"
    count = 0
    if method == "zscore":
        mu, sigma = valid.mean(), valid.std()
        if sigma == 0 or (sigma != sigma):
            return s, 0
        z = (s - mu) / sigma
        mask = np.abs(z) > 3
        count = mask.sum()
        low, high = mu - 3 * sigma, mu + 3 * sigma
        s = s.clip(lower=low, upper=high)
    else:
        q1, q3 = valid.quantile(0.25), valid.quantile(0.75)
        iqr = q3 - q1
        if iqr == 0 or (iqr != iqr):
            return s, 0
        low = q1 - 1.5 * iqr
        high = q3 + 1.5 * iqr
        count = ((s < low) | (s > high)).sum()
        s = s.clip(lower=low, upper=high)
    return s, int(count)


def apply_winsorization(df: pd.DataFrame) -> Tuple[pd.DataFrame, dict]:
    """
    对数值列应用盖帽法。成绩类用 Z-Score，频次/流量类用 IQR，其余按偏度自动选择。
    返回 (处理后的 df, {列名: 修正个数})
    """
    df = df.copy()
    stats_dict = {}
    for col in df.select_dtypes(include=[np.number]).columns:
        if col in ID_COL_NAMES or df[col].dtype == "object":
            continue
        if is_score_column(col):
            method = "zscore"
        elif is_duration_column(col) or "count" in str(col).lower() or "次数" in str(col) or "频" in str(col):
            method = "iqr"
        else:
            method = "auto"
        df[col], cnt = winsorize_column(df[col], method)
        if cnt > 0:
            stats_dict[col] = cnt
    return df, stats_dict


def impute_missing(df: pd.DataFrame) -> pd.DataFrame:
    """
    任务4：缺失值填补。
    - 连续数值型：中位数填补
    - 分类型：众数填补
    """
    df = df.copy()
    for col in df.columns:
        if df[col].isna().sum() == 0:
            continue
        if df[col].dtype in [np.float64, np.float32, np.int64, np.int32]:
            med = df[col].median()
            df[col] = df[col].fillna(med if not (pd.isna(med) or np.isnan(med)) else 0)
        else:
            mode_val = df[col].mode()
            if len(mode_val) > 0:
                df[col] = df[col].fillna(mode_val.iloc[0])
    return df


def clean_table(file_path: Path, field_mapping: dict) -> Optional[tuple[pd.DataFrame, dict]]:
    """对单表执行完整清洗流程。"""
    df = load_excel(file_path)
    if df is None and DATA_FALLBACK.exists():
        df = load_excel(DATA_FALLBACK / file_path.name)
    if df is None:
        return None
    tname = file_path.stem
    df = ensure_student_id(df, tname, field_mapping)
    df = hard_boundary_clipping(df)
    df, win_stats = apply_winsorization(df)
    df = impute_missing(df)
    return df, win_stats


def main():
    print("=" * 70)
    print("01_univariate_cleaning - 单变量异常值清洗")
    print("=" * 70)

    if not DATA_DIR.exists():
        print(f"[错误] 数据目录不存在: {DATA_DIR}")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    field_mapping = load_field_descriptions()
    files = sorted(DATA_DIR.glob("*.xlsx"))
    if not files and DATA_FALLBACK.exists():
        files = sorted(DATA_FALLBACK.glob("*.xlsx"))

    total_win = 0
    copied_failed = []
    for fp in files:
        result = clean_table(fp, field_mapping)
        if result is None:
            # 读取失败：将原文件复制到输出目录，保证 23 个文件齐全
            out_path = OUTPUT_DIR / fp.name
            try:
                shutil.copy2(fp, out_path)
                copied_failed.append(fp.name)
                print(f"[读取失败-已复制原文件] {fp.name}")
            except Exception as e:
                print(f"[跳过] {fp.name}: 读取失败且复制失败 - {e}")
            continue
        df, win_stats = result
        out_path = OUTPUT_DIR / fp.name
        try:
            df.to_excel(out_path, index=False, engine="openpyxl")
        except Exception as e:
            print(f"[错误] {fp.name}: 保存失败 - {e}")
            continue
        n_win = sum(win_stats.values())
        total_win += n_win
        if win_stats:
            print(f"[{fp.name}] 统计学异常值修正: {n_win} 个")
            for col, cnt in sorted(win_stats.items(), key=lambda x: -x[1])[:5]:
                print(f"    - {col}: {cnt}")
        else:
            print(f"[{fp.name}] 无统计学异常值修正")

    print("\n" + "=" * 70)
    print(f"清洗完成，输出至: {OUTPUT_DIR}")
    print(f"统计学异常值修正总计: {total_win} 个")
    if copied_failed:
        print(f"[提示] 以下文件因 ZIP 损坏无法解析，已复制原文件: {copied_failed}")
        print("       建议用 Excel 打开后另存为新文件，或从原始数据源重新导出")
    print("=" * 70)


if __name__ == "__main__":
    main()
